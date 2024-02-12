import sys
import os
import unittest
import datetime
from unittest.mock import patch
from openpyxl import Workbook, load_workbook

# Add the project root to the sys.path
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)

# Now you can import your modules from the project
from app.services.zerodha_service import login_in_zerodha
from app.services.fyers_auth_service import generate_access_token

zerodha_response = None
fyers_response = None


def set_zerodha_response(response):
    global zerodha_response
    zerodha_response = response


def set_fyers_response(response):
    global fyers_response
    fyers_response = response


class TestLogin(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Try to load existing workbook or create a new one
        try:
            cls.workbook = load_workbook("Login_UnitTest.xlsx")
        except FileNotFoundError:
            cls.workbook = Workbook()

        cls.sheet = cls.workbook.active
        cls.sheet.title = "Login Results"

        # If the workbook is newly created, write the header row
        if cls.sheet.max_row == 1:
            cls.sheet.append(
                [
                    "Zerodha Login Response",
                    "Fyers Login Response",
                    "Zerodha Passed",
                    "Fyers Passed",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(cls):
        cls.workbook.save("Login_UnitTest.xlsx")

    def write_to_excel(self):
        date = datetime.datetime.now().strftime("%Y-%m-%d")
        time = datetime.datetime.now().strftime("%H:%M:%S")
        zerodha_passed = zerodha_response is not None
        fyers_passed = fyers_response is not None

        self.sheet.append(
            [
                zerodha_response,
                fyers_response,
                zerodha_passed,
                fyers_passed,
                date,
                time,
            ]
        )

    @patch("app.services.zerodha_service.kite")
    @patch("app.services.fyers_auth_service.generate_access_token")
    def test_login_success(self, mock_fyers_api, mock_kite):
        mock_fyers_api.return_value = True
        mock_kite.return_value = True

        # Call the login functions
        result_zerodha = login_in_zerodha()
        result_fyers = generate_access_token()

        # Set responses
        set_zerodha_response(result_zerodha)
        set_fyers_response(result_fyers)

        # Write to Excel
        self.write_to_excel()


if __name__ == "__main__":
    unittest.main()
